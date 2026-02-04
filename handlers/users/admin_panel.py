from aiogram import types
from aiogram.dispatcher import FSMContext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import os
import tempfile
import json
import io

from loader import dp, db, bot
from keyboards.inline.buttons import get_admin_menu, get_stats_menu


async def is_admin(user_id: int) -> bool:
    return await db.is_admin(user_id)


@dp.message_handler(chat_type=types.ChatType.PRIVATE, commands=['admin'], state='*')
async def cmd_admin(message: types.Message, state: FSMContext):
    await state.finish()

    if not await is_admin(message.from_user.id):
        await message.answer("⛔ Sizda ruxsat yo'q!")
        return

    total_users = await db.count_users()
    total_surveys = await db.count_surveys()
    total_channels = await db.count_channels()
    pending_approvals = await db.count_pending_approvals()

    text = (
        "👨‍💼 <b>ADMIN PANEL</b>\n\n"
        f"👥 Jami foydalanuvchilar: <b>{total_users}</b>\n"
        f"⏳ Tasdiqlash kutayotganlar: <b>{pending_approvals}</b>\n"
        f"📋 So'rovnomalar: <b>{total_surveys}</b>\n"
        f"📢 Kanallar: <b>{total_channels}</b>\n\n"
        "Kerakli bo'limni tanlang:"
    )

    await message.answer(text, reply_markup=get_admin_menu())


@dp.callback_query_handler(text="admin:back", state='*')
async def callback_admin_back(callback: types.CallbackQuery, state: FSMContext):
    await state.finish()

    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    total_users = await db.count_users()
    total_surveys = await db.count_surveys()
    total_channels = await db.count_channels()
    pending_approvals = await db.count_pending_approvals()

    text = (
        "👨‍💼 <b>ADMIN PANEL</b>\n\n"
        f"👥 Jami foydalanuvchilar: <b>{total_users}</b>\n"
        f"⏳ Tasdiqlash kutayotganlar: <b>{pending_approvals}</b>\n"
        f"📋 So'rovnomalar: <b>{total_surveys}</b>\n"
        f"📢 Kanallar: <b>{total_channels}</b>\n\n"
        "Kerakli bo'limni tanlang:"
    )

    await callback.message.edit_text(text, reply_markup=get_admin_menu())
    await callback.answer()


@dp.callback_query_handler(text="admin:close", state='*')
async def callback_admin_close(callback: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await callback.message.delete()
    await callback.answer()


@dp.callback_query_handler(text="admin:stats", state='*')
async def callback_stats(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    total_users = await db.count_users()
    users_24h = await db.count_users_last_24h()
    users_week = await db.count_users_last_week()

    approved_users = await db.count_approved_users()
    pending_users = await db.count_pending_approvals()
    rejected_users = await db.count_rejected_users()

    active_survey = await db.get_active_survey()
    active_text = "Yo'q"
    responses_count = 0

    if active_survey:
        active_text = active_survey['name']
        responses_count = await db.count_survey_responses(active_survey['id'])

    text = (
        "📊 <b>STATISTIKA</b>\n\n"
        f"👥 <b>Foydalanuvchilar:</b>\n"
        f"   • Jami: <b>{total_users}</b>\n"
        f"   • Oxirgi 24 soat: <b>{users_24h}</b>\n"
        f"   • Oxirgi 7 kun: <b>{users_week}</b>\n\n"
        f"✅ <b>Tasdiqlangan:</b> <b>{approved_users}</b>\n"
        f"⏳ <b>Kutayotgan:</b> <b>{pending_users}</b>\n"
        f"❌ <b>Rad etilgan:</b> <b>{rejected_users}</b>\n\n"
        f"📋 <b>Aktiv so'rovnoma:</b> {active_text}\n"
        f"   • Javoblar soni: <b>{responses_count}</b>"
    )

    await callback.message.edit_text(text, reply_markup=get_stats_menu())
    await callback.answer()


@dp.callback_query_handler(text="stats:download", state='*')
async def callback_download_stats(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    # ✅ BIRINCHI callback.answer() - 30 sekund muammosi yo'q!
    await callback.answer()

    active_survey = await db.get_active_survey()

    if not active_survey:
        await callback.message.answer("⚠️ Aktiv so'rovnoma yo'q!")
        return

    fields = await db.get_survey_fields(active_survey['id'])
    responses = await db.get_survey_responses(active_survey['id'])

    if not responses:
        await callback.message.answer("⚠️ Javoblar yo'q!")
        return

    # Excel tayyorlanmoqda xabari
    await callback.message.edit_text("⏳ Excel fayl tayyorlanmoqda...")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Javoblar"

        headers = ["№"] + [f['column_name'] for f in fields] + ["Sana"]

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = border

        # Ma'lumotlar
        current_row = 2
        temp_images = []

        for response in responses:
            response_data = json.loads(response['response_data']) if isinstance(response['response_data'], str) else response['response_data']

            ws.cell(row=current_row, column=1, value=current_row - 1).border = border

            for col_idx, field in enumerate(fields, 2):
                value = response_data.get(field['column_name'], "")
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border

                # Rasm uchun
                if field['field_type'] == 'photo' and value:
                    try:
                        file = await bot.get_file(value)
                        file_path = file.file_path
                        downloaded_file = await bot.download_file(file_path)

                        img = Image.open(io.BytesIO(downloaded_file.read()))

                        if img.width < 300 or img.height < 300:
                            img = img.resize((300, 300), Image.Resampling.LANCZOS)
                        else:
                            img.thumbnail((300, 300), Image.Resampling.LANCZOS)

                        temp_img_path = os.path.join(tempfile.gettempdir(), f"temp_img_{current_row}_{col_idx}.png")
                        img.save(temp_img_path, "PNG", quality=100, optimize=False)
                        temp_images.append(temp_img_path)

                        xl_img = XLImage(temp_img_path)
                        xl_img.width = 300
                        xl_img.height = 300

                        ws.row_dimensions[current_row].height = 230
                        col_letter = cell.column_letter
                        ws.column_dimensions[col_letter].width = 42

                        ws.add_image(xl_img, f"{col_letter}{current_row}")
                        cell.value = ""

                    except Exception as e:
                        cell.value = "📷 Xato"
                        print(f"Rasm yuklashda xato: {e}")

                # Lokatsiya uchun
                elif field['field_type'] == 'location' and value:
                    try:
                        loc = json.loads(value) if isinstance(value, str) else value
                        link = f"https://maps.google.com/?q={loc['latitude']},{loc['longitude']}"
                        cell.value = link
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                    except:
                        cell.value = "Lokatsiya"

                else:
                    cell.value = value
                    cell.alignment = Alignment(wrap_text=True, vertical='center')

            # Sana
            cell = ws.cell(row=current_row, column=len(headers), value=str(response['submitted_at']))
            cell.border = border

            current_row += 1

        # Oddiy ustunlar uchun kenglik
        for col in range(1, len(headers) + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            if ws.column_dimensions[col_letter].width == 13.0:
                ws.column_dimensions[col_letter].width = 20

        file_path = os.path.join(tempfile.gettempdir(), active_survey['file_name'])
        wb.save(file_path)

        with open(file_path, 'rb') as file:
            await bot.send_document(
                callback.from_user.id,
                file,
                caption=f"📊 <b>{active_survey['name']}</b>\n👥 Jami javoblar: {len(responses)}"
            )

        # Temp fayllarni o'chirish
        os.remove(file_path)
        for temp_img in temp_images:
            try:
                os.remove(temp_img)
            except:
                pass

        await callback.message.edit_text(
            "✅ Excel fayl yuborildi!",
            reply_markup=get_stats_menu()
        )

    except Exception as e:
        print(f"Excel yaratishda xato: {e}")
        await callback.message.edit_text(
            "❌ Xatolik yuz berdi!",
            reply_markup=get_stats_menu()
        )