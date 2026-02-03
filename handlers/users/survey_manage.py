from aiogram import types
from aiogram.dispatcher import FSMContext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import os
import json
import tempfile
import io

from loader import dp, db, bot
from keyboards.inline.buttons import (
    get_survey_list_keyboard, get_survey_actions,
    get_survey_delete_confirm, get_surveys_menu
)
from handlers.users.admin_panel import is_admin


@dp.callback_query_handler(text="survey:list", state='*')
async def callback_survey_list(callback: types.CallbackQuery, state: FSMContext):
    await state.finish()

    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    surveys = await db.get_all_surveys()

    if not surveys:
        await callback.message.edit_text(
            "📋 <b>So'rovnomalar ro'yxati</b>\n\n"
            "Hozircha so'rovnomalar yo'q.",
            reply_markup=get_surveys_menu()
        )
        await callback.answer()
        return

    text = "📋 <b>So'rovnomalar ro'yxati</b>\n\n"
    text += "✅ - Aktiv | ⏸ - Noaktiv\n\n"

    await callback.message.edit_text(text, reply_markup=get_survey_list_keyboard(surveys))
    await callback.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("survey:view:"), state='*')
async def callback_view_survey(callback: types.CallbackQuery, state: FSMContext):
    await state.finish()

    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    survey_id = int(callback.data.split(":")[2])
    survey = await db.get_survey(survey_id)

    if not survey:
        await callback.answer("So'rovnoma topilmadi!", show_alert=True)
        return

    fields = await db.get_survey_fields(survey_id)
    responses_count = await db.count_survey_responses(survey_id)

    status = "✅ Aktiv" if survey['is_active'] else "⏸ Noaktiv"

    text = (
        f"📋 <b>{survey['name']}</b>\n\n"
        f"📁 Fayl: <code>{survey['file_name']}</code>\n"
        f"📊 Holat: {status}\n"
        f"👥 Javoblar: <b>{responses_count}</b> ta\n\n"
        f"<b>Ustunlar ({len(fields)} ta):</b>\n"
    )

    field_types = {
        'text': '📝',
        'choice': '🔘',
        'photo': '📷',
        'location': '📍'
    }

    for i, field in enumerate(fields, 1):
        icon = field_types.get(field['field_type'], '❓')
        if field['field_type'] == 'choice':
            opts = field['options'] if field['options'] else []
            text += f"{i}. {field['column_name']} ({icon} {len(opts)})\n"
        else:
            text += f"{i}. {field['column_name']} ({icon})\n"

    await callback.message.edit_text(text, reply_markup=get_survey_actions(survey_id, survey['is_active']))
    await callback.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("survey:activate:"), state='*')
async def callback_activate_survey(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    survey_id = int(callback.data.split(":")[2])
    await db.set_survey_active(survey_id)

    survey = await db.get_survey(survey_id)
    fields = await db.get_survey_fields(survey_id)
    responses_count = await db.count_survey_responses(survey_id)

    text = (
        f"📋 <b>{survey['name']}</b>\n\n"
        f"📁 Fayl: <code>{survey['file_name']}</code>\n"
        f"📊 Holat: ✅ Aktiv\n"
        f"👥 Javoblar: <b>{responses_count}</b> ta\n\n"
        f"<b>Ustunlar ({len(fields)} ta):</b>\n"
    )

    field_types = {
        'text': '📝',
        'choice': '🔘',
        'photo': '📷',
        'location': '📍'
    }

    for i, field in enumerate(fields, 1):
        icon = field_types.get(field['field_type'], '❓')
        if field['field_type'] == 'choice':
            opts = field['options'] if field['options'] else []
            text += f"{i}. {field['column_name']} ({icon} {len(opts)})\n"
        else:
            text += f"{i}. {field['column_name']} ({icon})\n"

    await callback.message.edit_text(text, reply_markup=get_survey_actions(survey_id, True))
    await callback.answer(f"✅ {survey['name']} aktiv qilindi!", show_alert=True)


@dp.callback_query_handler(lambda c: c.data.startswith("survey:deactivate:"), state='*')
async def callback_deactivate_survey(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    survey_id = int(callback.data.split(":")[2])
    await db.deactivate_survey(survey_id)

    survey = await db.get_survey(survey_id)
    fields = await db.get_survey_fields(survey_id)
    responses_count = await db.count_survey_responses(survey_id)

    text = (
        f"📋 <b>{survey['name']}</b>\n\n"
        f"📁 Fayl: <code>{survey['file_name']}</code>\n"
        f"📊 Holat: ⏸ Noaktiv\n"
        f"👥 Javoblar: <b>{responses_count}</b> ta\n\n"
        f"<b>Ustunlar ({len(fields)} ta):</b>\n"
    )

    field_types = {
        'text': '📝',
        'choice': '🔘',
        'photo': '📷',
        'location': '📍'
    }

    for i, field in enumerate(fields, 1):
        icon = field_types.get(field['field_type'], '❓')
        if field['field_type'] == 'choice':
            opts = field['options'] if field['options'] else []
            text += f"{i}. {field['column_name']} ({icon} {len(opts)})\n"
        else:
            text += f"{i}. {field['column_name']} ({icon})\n"

    await callback.message.edit_text(text, reply_markup=get_survey_actions(survey_id, False))
    await callback.answer(f"⏸ {survey['name']} deaktiv qilindi!", show_alert=True)


@dp.callback_query_handler(lambda c: c.data.startswith("survey:excel:"), state='*')
async def callback_download_survey_excel(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    survey_id = int(callback.data.split(":")[2])
    survey = await db.get_survey(survey_id)

    if not survey:
        await callback.answer("So'rovnoma topilmadi!", show_alert=True)
        return

    fields = await db.get_survey_fields(survey_id)
    responses = await db.get_survey_responses(survey_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Javoblar"

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["№"] + [f['column_name'] for f in fields] + ["Sana"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = border

    current_row = 2
    temp_images = []

    for response in responses:
        response_data = json.loads(response['response_data']) if isinstance(response['response_data'], str) else response['response_data']

        ws.cell(row=current_row, column=1, value=current_row - 1).border = border

        for col_idx, field in enumerate(fields, 2):
            value = response_data.get(field['column_name'], "")
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border

            # Rasm uchun
            if field['field_type'] == 'photo' and value:
                try:
                    # Rasmni yuklab olish
                    file = await bot.get_file(value)
                    file_path = file.file_path
                    downloaded_file = await bot.download_file(file_path)

                    # Rasmni qayta o'lchamlash
                    img = Image.open(io.BytesIO(downloaded_file.read()))

                    # Bir xil o'lchamga keltirish (150x150)
                    img.thumbnail((150, 150), Image.Resampling.LANCZOS)

                    # Temp faylga saqlash
                    temp_img_path = os.path.join(tempfile.gettempdir(), f"temp_img_{current_row}_{col_idx}.png")
                    img.save(temp_img_path, "PNG")
                    temp_images.append(temp_img_path)

                    # Excel'ga qo'shish
                    xl_img = XLImage(temp_img_path)
                    xl_img.width = 150
                    xl_img.height = 150

                    # Katakchani kattalashtirish
                    ws.row_dimensions[current_row].height = 115
                    col_letter = ws.cell(row=current_row, column=col_idx).column_letter
                    ws.column_dimensions[col_letter].width = 22

                    ws.add_image(xl_img, f"{col_letter}{current_row}")
                    cell.value = ""

                except Exception as e:
                    cell.value = "📷 Xato"
                    print(f"Rasm yuklashda xato: {e}")

            # Lokatsiya uchun
            elif field['field_type'] == 'location' and value:
                try:
                    loc = json.loads(value) if isinstance(value, str) else value
                    link = f"https://maps.google.com/?q={loc['latitude']},{loc['longitude']}"
                    cell.value = link
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                except:
                    cell.value = "Lokatsiya"

            else:
                cell.value = value
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Sana
        cell = ws.cell(row=current_row, column=len(headers), value=str(response['submitted_at']))
        cell.border = border

        current_row += 1

    # Oddiy ustunlar uchun kenglik
    for col in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        if ws.column_dimensions[col_letter].width == 13.0:
            ws.column_dimensions[col_letter].width = 20

    file_path = os.path.join(tempfile.gettempdir(), survey['file_name'])
    wb.save(file_path)

    with open(file_path, 'rb') as file:
        await callback.message.answer_document(
            file,
            caption=f"📊 <b>{survey['name']}</b>\n👥 Javoblar: {len(responses)} ta"
        )

    # Temp fayllarni o'chirish
    os.remove(file_path)
    for temp_img in temp_images:
        try:
            os.remove(temp_img)
        except:
            pass

    await callback.answer("Yuklandi!")


@dp.callback_query_handler(lambda c: c.data.startswith("survey:delete:") and "confirm" not in c.data, state='*')
async def callback_delete_survey(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    survey_id = int(callback.data.split(":")[2])
    survey = await db.get_survey(survey_id)

    if not survey:
        await callback.answer("So'rovnoma topilmadi!", show_alert=True)
        return

    responses_count = await db.count_survey_responses(survey_id)

    await callback.message.edit_text(
        f"🗑 <b>So'rovnomani o'chirish</b>\n\n"
        f"📋 <b>{survey['name']}</b>\n"
        f"👥 Javoblar: {responses_count} ta\n\n"
        f"⚠️ Barcha javoblar ham o'chib ketadi!\n\n"
        f"Tasdiqlaysizmi?",
        reply_markup=get_survey_delete_confirm(survey_id)
    )
    await callback.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("survey:delete_confirm:"), state='*')
async def callback_delete_survey_confirm(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    survey_id = int(callback.data.split(":")[2])
    survey = await db.get_survey(survey_id)

    if not survey:
        await callback.answer("So'rovnoma topilmadi!", show_alert=True)
        return

    await db.delete_survey(survey_id)

    await callback.message.edit_text(f"✅ <b>{survey['name']}</b> o'chirildi!")
    await callback.answer("O'chirildi!")


@dp.callback_query_handler(lambda c: c.data.startswith("survey:edit:"), state='*')
async def callback_edit_survey(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    await callback.answer("🔜 Tez orada qo'shiladi!", show_alert=True)