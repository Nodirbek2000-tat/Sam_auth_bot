from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup

from loader import dp, bot
from states.register import Registration
from keyboards.inline.buttons import get_yunalish_keyboard, confirm_keyboard, YUNALISHLAR

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
import os
import json

# ============ Sozlamalar ============
EXCEL_FILE = "royxatlar__.xlsx"
ADMINS_FILE = "admins.json"
SUPER_ADMIN = 736290914  # Asosiy admin

# ============ KANALLAR RO'YXATI ============
# Yangi kanal qo'shish uchun # belgisini olib tashla va ma'lumotlarni yoz

CHANNELS = [
    # 1-kanal: https://t.me/Samstf
    ("StartUplar Ofisi", "Samstf"),

    # 2-kanal: https://t.me/yiasamyosh
    ("SAMARQANDLIK YOSHLAR", "yiasamyosh"),

    # 3-kanal: https://t.me/...
    # ("Kanal nomi", "kanal_username"),

    # 4-kanal: https://t.me/...
    # ("Kanal nomi", "kanal_username"),

    # 5-kanal: https://t.me/...
    # ("Kanal nomi", "kanal_username"),
]


# ============ OBUNA TEKSHIRISH ============

async def check_subscription(user_id: int) -> dict:
    """Foydalanuvchi barcha kanallarga obuna bo'lganini tekshirish"""
    not_subscribed = []

    for channel_name, channel_username in CHANNELS:
        try:
            member = await bot.get_chat_member(
                chat_id=f"@{channel_username}",
                user_id=user_id
            )

            if member.status in ["left", "kicked"]:
                not_subscribed.append({
                    "name": channel_name,
                    "username": channel_username
                })

        except Exception as e:
            print(f"Kanal tekshirishda xatolik ({channel_username}): {e}")
            not_subscribed.append({
                "name": channel_name,
                "username": channel_username
            })

    return {
        "is_subscribed": len(not_subscribed) == 0,
        "not_subscribed": not_subscribed
    }


def get_subscribe_keyboard(not_subscribed: list) -> InlineKeyboardMarkup:
    """Obuna bo'lish tugmalarini yaratish"""
    keyboard = InlineKeyboardMarkup(row_width=1)

    for channel in not_subscribed:
        keyboard.add(
            InlineKeyboardButton(
                text=f"📢 {channel['name']}",
                url=f"https://t.me/{channel['username']}"
            )
        )

    keyboard.add(
        InlineKeyboardButton(
            text="✅ Tekshirish",
            callback_data="check_subscription"
        )
    )

    return keyboard


async def check_and_request_subscription(message: types.Message) -> bool:
    """Obunani tekshirish"""
    result = await check_subscription(message.from_user.id)

    if not result["is_subscribed"]:
        await message.answer(
            "⚠️ <b>Botdan foydalanish uchun quyidagi kanallarga obuna bo'ling:</b>",
            reply_markup=get_subscribe_keyboard(result["not_subscribed"])
        )
        return False

    return True


@dp.callback_query_handler(text="check_subscription", state="*")
async def callback_check_subscription(callback: types.CallbackQuery, state: FSMContext):
    """Obunani qayta tekshirish"""
    result = await check_subscription(callback.from_user.id)

    if result["is_subscribed"]:
        await callback.message.edit_text(
            "✅ <b>Rahmat! Siz barcha kanallarga obuna bo'ldingiz.</b>\n\n"
            "📝 Ro'yxatdan o'tish uchun /register buyrug'ini bosing."
        )
    else:
        await callback.message.edit_text(
            "⚠️ <b>Siz hali barcha kanallarga obuna bo'lmadingiz!</b>\n\n"
            "Quyidagi kanallarga obuna bo'ling:",
            reply_markup=get_subscribe_keyboard(result["not_subscribed"])
        )

    await callback.answer()


# ============ Admin funksiyalari ============
class AdminStates(StatesGroup):
    add_admin = State()
    remove_admin = State()


def load_admins():
    """Adminlarni yuklash"""
    if os.path.exists(ADMINS_FILE):
        with open(ADMINS_FILE, 'r') as f:
            return json.load(f)
    return [SUPER_ADMIN]


def save_admins(admins: list):
    """Adminlarni saqlash"""
    with open(ADMINS_FILE, 'w') as f:
        json.dump(admins, f)


def is_admin(user_id: int) -> bool:
    """Admin ekanligini tekshirish"""
    admins = load_admins()
    return user_id in admins or user_id == SUPER_ADMIN


def is_super_admin(user_id: int) -> bool:
    """Super admin ekanligini tekshirish"""
    return user_id == SUPER_ADMIN


# ============ Excel funksiyalari ============
def get_border():
    """Border style"""
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def create_excel_if_not_exists():
    """Excel fayl yaratish (agar yo'q bo'lsa)"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Ro'yxat"

        headers = [
            "№",
            "Фамилия, исми, шарифи",
            "Туғилган куни, ойи, йили",
            "Яшаш манзили (шаҳар ёки туман номи, маҳалласи, кўчаси, хонадон рақами)",
            "Ўқиш жойи",
            "Телефон рақами",
            "Танлаган клубдаги йўналиш номи",
            "Қизиқишлари ва эришган ютуқлари"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = get_border()

        column_widths = [5, 25, 20, 45, 25, 18, 35, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        wb.save(EXCEL_FILE)


def save_to_excel(data: dict):
    """Ma'lumotlarni Excel'ga saqlash"""
    try:
        create_excel_if_not_exists()

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        last_row = ws.max_row + 1
        nomer = last_row - 1

        row_data = [
            nomer,
            data.get("fio", ""),
            data.get("tug_sana", ""),
            data.get("manzil", ""),
            data.get("oquv_joy", ""),
            data.get("telefon", ""),
            data.get("yunalish", ""),
            data.get("qiziqish", "")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=last_row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = get_border()

        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Excel xatosi: {e}")
        return False


def get_statistics():
    """Statistika olish"""
    if not os.path.exists(EXCEL_FILE):
        return None

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    total = ws.max_row - 1
    return total if total > 0 else 0


# ============ User Handlers ============

@dp.message_handler(commands=['start'])
async def cmd_start(message: types.Message):
    # OBUNANI TEKSHIRISH
    if not await check_and_request_subscription(message):
        return

    await message.answer(
        "👋 Assalomu alaykum!\n\n"
        "📝 Ro'yxatdan o'tish uchun /register buyrug'ini bosing."
    )


@dp.message_handler(commands=['register'])
async def cmd_register(message: types.Message):
    # OBUNANI TEKSHIRISH
    if not await check_and_request_subscription(message):
        return

    await message.answer(
        "📝 <b>Ro'yxatdan o'tish</b>\n\n"
        "✏️ Familiya, ism, sharifingizni kiriting:\n"
        "<i>(Masalan: Karimov Alisher Nodirovich)</i>"
    )
    await Registration.fio.set()


@dp.message_handler(state=Registration.fio)
async def process_fio(message: types.Message, state: FSMContext):
    await state.update_data(fio=message.text)
    await message.answer(
        "📅 Tug'ilgan kuningizni kiriting:\n"
        "<i>(Masalan: 15.03.2000)</i>"
    )
    await Registration.tug_sana.set()


@dp.message_handler(state=Registration.tug_sana)
async def process_tug_sana(message: types.Message, state: FSMContext):
    await state.update_data(tug_sana=message.text)
    await message.answer(
        "📍 Yashash manzilingizni kiriting:\n"
        "<i>(Shahar/tuman, mahalla, ko'cha, uy raqami)</i>"
    )
    await Registration.manzil.set()


@dp.message_handler(state=Registration.manzil)
async def process_manzil(message: types.Message, state: FSMContext):
    await state.update_data(manzil=message.text)
    await message.answer(
        "🎓 O'qish joyingizni kiriting:\n"
        "<i>(Masalan: TATU, 2-kurs)</i>"
    )
    await Registration.oquv_joy.set()


@dp.message_handler(state=Registration.oquv_joy)
async def process_oquv_joy(message: types.Message, state: FSMContext):
    await state.update_data(oquv_joy=message.text)
    await message.answer(
        "📱 Telefon raqamingizni kiriting:\n"
        "<i>(Masalan: +998901234567)</i>"
    )
    await Registration.telefon.set()


@dp.message_handler(state=Registration.telefon)
async def process_telefon(message: types.Message, state: FSMContext):
    await state.update_data(telefon=message.text)
    await message.answer(
        "📋 Yo'nalishni tanlang:",
        reply_markup=get_yunalish_keyboard()
    )
    await Registration.yunalish.set()


@dp.callback_query_handler(lambda c: c.data.startswith("yunalish_"), state=Registration.yunalish)
async def process_yunalish(callback: types.CallbackQuery, state: FSMContext):
    yunalish_index = int(callback.data.split("_")[1]) - 1
    yunalish_name = YUNALISHLAR[yunalish_index]

    await state.update_data(yunalish=yunalish_name)

    await callback.message.edit_text(
        "🎯 Qiziqishlaringiz va erishgan yutuqlaringizni yozing:"
    )
    await Registration.qiziqish.set()


@dp.message_handler(state=Registration.qiziqish)
async def process_qiziqish(message: types.Message, state: FSMContext):
    await state.update_data(qiziqish=message.text)

    data = await state.get_data()

    text = (
        "📋 <b>Ma'lumotlaringizni tekshiring:</b>\n\n"
        f"👤 F.I.Sh: {data['fio']}\n"
        f"📅 Tug'ilgan sana: {data['tug_sana']}\n"
        f"📍 Manzil: {data['manzil']}\n"
        f"🎓 O'qish joyi: {data['oquv_joy']}\n"
        f"📱 Telefon: {data['telefon']}\n"
        f"📋 Yo'nalish: {data['yunalish']}\n"
        f"🎯 Qiziqish/Yutuqlar: {data['qiziqish']}\n\n"
        "✅ Tasdiqlaysizmi?"
    )

    await message.answer(text, reply_markup=confirm_keyboard)
    await Registration.tasdiqlash.set()


@dp.callback_query_handler(text="confirm_yes", state=Registration.tasdiqlash)
async def confirm_yes(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()

    if save_to_excel(data):
        await callback.message.edit_text(
            "✅ <b>Tabriklaymiz!</b>\n\n"
            "Siz muvaffaqiyatli ro'yxatdan o'tdingiz! 🎉"
        )
    else:
        await callback.message.edit_text(
            "❌ Xatolik yuz berdi.\n"
            "Iltimos, keyinroq qayta urinib ko'ring."
        )

    await state.finish()


@dp.callback_query_handler(text="confirm_no", state=Registration.tasdiqlash)
async def confirm_no(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "❌ Bekor qilindi.\n\n"
        "Qaytadan boshlash uchun /register buyrug'ini bosing."
    )
    await state.finish()


@dp.message_handler(commands=['cancel'], state='*')
async def cmd_cancel(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is None:
        return

    await state.finish()
    await message.answer("❌ Bekor qilindi.\n\nQaytadan: /register")


# ============ ADMIN PANEL ============

@dp.message_handler(commands=['admins'])
async def cmd_admins(message: types.Message):
    """Admin panel"""
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Sizda ruxsat yo'q!")
        return

    total = get_statistics()
    admins = load_admins()

    text = (
        "👨‍💼 <b>ADMIN PANEL</b>\n\n"
        f"📊 Ro'yxatdan o'tganlar: <b>{total if total else 0}</b> ta\n"
        f"👥 Adminlar soni: <b>{len(admins)}</b> ta\n\n"
        "<b>Buyruqlar:</b>\n"
        "/statistika - Excel faylni yuklash\n"
        "/adminlar - Adminlar ro'yxati\n"
    )

    if is_super_admin(message.from_user.id):
        text += (
            "/addadmin - Admin qo'shish\n"
            "/removeadmin - Adminni o'chirish\n"
        )

    await message.answer(text)


@dp.message_handler(commands=['statistika'])
async def cmd_statistika(message: types.Message):
    """Admin uchun Excel yuborish"""
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Sizda ruxsat yo'q!")
        return

    if not os.path.exists(EXCEL_FILE):
        await message.answer("📭 Hali hech kim ro'yxatdan o'tmagan.")
        return

    total = get_statistics()

    with open(EXCEL_FILE, 'rb') as file:
        await message.answer_document(
            file,
            caption=f"📊 Ro'yxatdan o'tganlar jadvali\n👥 Jami: {total} ta"
        )


@dp.message_handler(commands=['adminlar'])
async def cmd_adminlar(message: types.Message):
    """Adminlar ro'yxati"""
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Sizda ruxsat yo'q!")
        return

    admins = load_admins()

    text = "👥 <b>Adminlar ro'yxati:</b>\n\n"
    for i, admin_id in enumerate(admins, 1):
        if admin_id == SUPER_ADMIN:
            text += f"{i}. <code>{admin_id}</code> 👑 (Super Admin)\n"
        else:
            text += f"{i}. <code>{admin_id}</code>\n"

    await message.answer(text)


@dp.message_handler(commands=['addadmin'])
async def cmd_addadmin(message: types.Message):
    """Admin qo'shish"""
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Faqat Super Admin admin qo'sha oladi!")
        return

    await message.answer(
        "👤 Yangi admin Telegram ID sini yuboring:\n\n"
        "<i>ID bilish uchun: @userinfobot</i>"
    )
    await AdminStates.add_admin.set()


@dp.message_handler(state=AdminStates.add_admin)
async def process_add_admin(message: types.Message, state: FSMContext):
    """Adminni qo'shish"""
    try:
        new_admin_id = int(message.text)
    except ValueError:
        await message.answer("❌ Noto'g'ri ID! Faqat raqam kiriting.")
        return

    admins = load_admins()

    if new_admin_id in admins:
        await message.answer("⚠️ Bu foydalanuvchi allaqachon admin!")
        await state.finish()
        return

    admins.append(new_admin_id)
    save_admins(admins)

    await message.answer(
        f"✅ Yangi admin qo'shildi!\n\n"
        f"🆔 ID: <code>{new_admin_id}</code>"
    )
    await state.finish()


@dp.message_handler(commands=['removeadmin'])
async def cmd_removeadmin(message: types.Message):
    """Adminni o'chirish"""
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Faqat Super Admin adminni o'chira oladi!")
        return

    admins = load_admins()

    text = "👥 <b>Adminlar:</b>\n\n"
    for i, admin_id in enumerate(admins, 1):
        if admin_id == SUPER_ADMIN:
            text += f"{i}. <code>{admin_id}</code> 👑 (o'chirib bo'lmaydi)\n"
        else:
            text += f"{i}. <code>{admin_id}</code>\n"

    text += "\n❌ O'chirmoqchi bo'lgan admin ID sini yuboring:"

    await message.answer(text)
    await AdminStates.remove_admin.set()


@dp.message_handler(state=AdminStates.remove_admin)
async def process_remove_admin(message: types.Message, state: FSMContext):
    """Adminni o'chirish"""
    try:
        admin_id = int(message.text)
    except ValueError:
        await message.answer("❌ Noto'g'ri ID! Faqat raqam kiriting.")
        return

    if admin_id == SUPER_ADMIN:
        await message.answer("⛔ Super Adminni o'chirib bo'lmaydi!")
        await state.finish()
        return

    admins = load_admins()

    if admin_id not in admins:
        await message.answer("⚠️ Bu foydalanuvchi admin emas!")
        await state.finish()
        return

    admins.remove(admin_id)
    save_admins(admins)

    await message.answer(
        f"✅ Admin o'chirildi!\n\n"
        f"🆔 ID: <code>{admin_id}</code>"
    )
    await state.finish()